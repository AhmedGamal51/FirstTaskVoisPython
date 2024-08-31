import openpyxl
from datetime import datetime

def calculate_years_in_vois(joining_date):
    """
    Calculate the number of years between the joining date and the current date.
    """
    current_date = datetime.now()
    # Calculate the difference in years
    return current_date.year - joining_date.year - ((current_date.month, current_date.day) < (joining_date.month, joining_date.day))

# Define the path to your Excel file
file_path = r'C:\\Users\\Dell\\Desktop\\Vois_Task_Python\\TaskData.xlsx'  # Replace with your actual file path

# Load the workbook and select the active sheet
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

# Iterate over the rows in the sheet starting from the second row (to skip header)
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
    join_date_cell = row[2]  # Assuming "Join Date" is in the 1st column (index 0)
    years_in_vois_cell = row[3]  # Assuming "How Many years in VOIS" is in the 3rd column (index 2)

    if join_date_cell.value:
        # Check if the value is already a datetime object
        if isinstance(join_date_cell.value, datetime):
            join_date = join_date_cell.value
        else:
            try:
                # Attempt to parse the date from the specific format
                join_date = datetime.strptime(join_date_cell.value, "%A, %B %d, %Y")  # Adjusted format to match "Sunday, January 10, 2021"
            except (ValueError, TypeError):
                print(f"Row {row[0].row}: Invalid format for Joining Date: {join_date_cell.value}")
                continue  # Skip rows with invalid or unrecognized date format

        # Calculate the number of years since the joining date
        years_in_vois = calculate_years_in_vois(join_date)
        years_in_vois_cell.value = years_in_vois
        print(f"Row {row[0].row}: Joining Date = {join_date}, Years in _VOIS = {years_in_vois}")
    else:
        print(f"Row {row[0].row}: Invalid or empty Joining Date")  # Debug line for invalid or empty cells

# Save the updated Excel file
workbook.save(file_path)  # Make sure to save the file after making changes
print("Excel file updated successfully!")

